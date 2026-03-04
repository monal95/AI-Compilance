"""
Reporting Service - Generates compliance reports and analytics.

Features:
- PDF report generation
- Excel export
- Dashboard analytics
- Trend analysis
- Scheduled reporting
"""
from fastapi import FastAPI, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timedelta
from enum import Enum
import logging
import io
import json

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Reporting Service",
    description="Compliance report generation and analytics",
    version="1.0.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Enums
class ReportFormat(str, Enum):
    PDF = "pdf"
    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"


class ReportType(str, Enum):
    SINGLE_AUDIT = "single_audit"
    BULK_AUDIT = "bulk_audit"
    SUMMARY = "summary"
    TREND = "trend"
    VIOLATION = "violation"


# Models
class ReportRequest(BaseModel):
    report_type: ReportType
    format: ReportFormat = ReportFormat.JSON
    audit_ids: Optional[list[str]] = None
    date_from: Optional[datetime] = None
    date_to: Optional[datetime] = None
    category: Optional[str] = None
    risk_level: Optional[str] = None


class DashboardMetrics(BaseModel):
    total_audits: int
    compliant_count: int
    non_compliant_count: int
    average_score: float
    critical_violations: int
    audits_today: int
    score_trend: list[dict]
    category_distribution: dict
    risk_distribution: dict
    top_violations: list[dict]


# Sample data store (would be MongoDB in production)
sample_audits = [
    {
        "audit_id": "audit_001",
        "product_name": "Health Drink 500ml",
        "category": "food",
        "compliance_score": 85.5,
        "risk_level": "low",
        "violations": [
            {"rule": "net_weight_format", "severity": "medium", "message": "Net weight unit unclear"}
        ],
        "created_at": datetime.utcnow() - timedelta(hours=2),
    },
    {
        "audit_id": "audit_002",
        "product_name": "Organic Biscuits",
        "category": "food",
        "compliance_score": 45.0,
        "risk_level": "high",
        "violations": [
            {"rule": "fssai_missing", "severity": "critical", "message": "FSSAI number missing"},
            {"rule": "expiry_missing", "severity": "critical", "message": "Expiry date not found"},
        ],
        "created_at": datetime.utcnow() - timedelta(hours=5),
    },
    {
        "audit_id": "audit_003",
        "product_name": "Instant Noodles",
        "category": "food",
        "compliance_score": 92.0,
        "risk_level": "compliant",
        "violations": [],
        "created_at": datetime.utcnow() - timedelta(days=1),
    },
]


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "service": "reporting-service"}


@app.get("/reports/dashboard", response_model=DashboardMetrics)
async def get_dashboard_metrics(
    days: int = Query(7, ge=1, le=90, description="Number of days to analyze"),
):
    """Get dashboard metrics and KPIs."""
    
    # Filter audits by date range
    cutoff = datetime.utcnow() - timedelta(days=days)
    audits = [a for a in sample_audits if a['created_at'] >= cutoff]
    
    total = len(audits)
    compliant = sum(1 for a in audits if a['risk_level'] == 'compliant')
    non_compliant = total - compliant
    
    scores = [a['compliance_score'] for a in audits]
    avg_score = sum(scores) / len(scores) if scores else 0
    
    # Critical violations count
    critical_violations = sum(
        1 for a in audits 
        for v in a.get('violations', [])
        if v.get('severity') == 'critical'
    )
    
    # Today's audits
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    audits_today = sum(1 for a in audits if a['created_at'] >= today_start)
    
    # Score trend (daily averages)
    score_trend = []
    for i in range(min(days, 7)):
        day = datetime.utcnow() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        day_audits = [a for a in audits if day_start <= a['created_at'] < day_end]
        if day_audits:
            day_avg = sum(a['compliance_score'] for a in day_audits) / len(day_audits)
        else:
            day_avg = 0
        
        score_trend.append({
            "date": day_start.strftime("%Y-%m-%d"),
            "average_score": round(day_avg, 1),
            "count": len(day_audits),
        })
    
    # Category distribution
    category_dist = {}
    for a in audits:
        cat = a.get('category', 'unknown')
        category_dist[cat] = category_dist.get(cat, 0) + 1
    
    # Risk distribution
    risk_dist = {}
    for a in audits:
        risk = a.get('risk_level', 'unknown')
        risk_dist[risk] = risk_dist.get(risk, 0) + 1
    
    # Top violations
    violation_counts = {}
    for a in audits:
        for v in a.get('violations', []):
            rule = v.get('rule', 'unknown')
            violation_counts[rule] = violation_counts.get(rule, 0) + 1
    
    top_violations = sorted(
        [{"rule": k, "count": v} for k, v in violation_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )[:10]
    
    return DashboardMetrics(
        total_audits=total,
        compliant_count=compliant,
        non_compliant_count=non_compliant,
        average_score=round(avg_score, 1),
        critical_violations=critical_violations,
        audits_today=audits_today,
        score_trend=score_trend,
        category_distribution=category_dist,
        risk_distribution=risk_dist,
        top_violations=top_violations,
    )


@app.post("/reports/generate")
async def generate_report(request: ReportRequest):
    """Generate a compliance report."""
    
    if request.format == ReportFormat.JSON:
        return await generate_json_report(request)
    elif request.format == ReportFormat.CSV:
        return await generate_csv_report(request)
    elif request.format == ReportFormat.PDF:
        return await generate_pdf_report(request)
    elif request.format == ReportFormat.EXCEL:
        return await generate_excel_report(request)
    else:
        raise HTTPException(status_code=400, detail="Unsupported format")


async def generate_json_report(request: ReportRequest):
    """Generate JSON format report."""
    audits = filter_audits(request)
    
    if request.report_type == ReportType.SINGLE_AUDIT:
        if not request.audit_ids:
            raise HTTPException(status_code=400, detail="audit_ids required")
        return {"audits": audits}
    
    elif request.report_type == ReportType.SUMMARY:
        return generate_summary(audits)
    
    elif request.report_type == ReportType.VIOLATION:
        return generate_violation_report(audits)
    
    return {"audits": audits}


async def generate_csv_report(request: ReportRequest):
    """Generate CSV format report."""
    audits = filter_audits(request)
    
    # Create CSV content
    lines = ["audit_id,product_name,category,compliance_score,risk_level,violation_count,created_at"]
    
    for audit in audits:
        lines.append(
            f"{audit['audit_id']},"
            f"\"{audit['product_name']}\","
            f"{audit['category']},"
            f"{audit['compliance_score']},"
            f"{audit['risk_level']},"
            f"{len(audit.get('violations', []))},"
            f"{audit['created_at'].isoformat()}"
        )
    
    content = "\n".join(lines)
    
    return StreamingResponse(
        io.StringIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=compliance_report.csv"}
    )


async def generate_pdf_report(request: ReportRequest):
    """Generate PDF format report."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        
        audits = filter_audits(request)
        buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
        )
        story.append(Paragraph("Compliance Audit Report", title_style))
        story.append(Spacer(1, 12))
        
        # Report info
        story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
        story.append(Paragraph(f"Total Audits: {len(audits)}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary stats
        if audits:
            avg_score = sum(a['compliance_score'] for a in audits) / len(audits)
            compliant = sum(1 for a in audits if a['risk_level'] == 'compliant')
            
            story.append(Paragraph("Summary", styles['Heading2']))
            story.append(Paragraph(f"Average Compliance Score: {avg_score:.1f}%", styles['Normal']))
            story.append(Paragraph(f"Compliant Products: {compliant}/{len(audits)}", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Audit table
        story.append(Paragraph("Audit Details", styles['Heading2']))
        
        table_data = [['Product', 'Category', 'Score', 'Risk', 'Violations']]
        for audit in audits[:20]:  # Limit to 20 for PDF
            table_data.append([
                audit['product_name'][:30],
                audit['category'],
                f"{audit['compliance_score']:.1f}%",
                audit['risk_level'],
                str(len(audit.get('violations', [])))
            ])
        
        table = Table(table_data, colWidths=[2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=compliance_report.pdf"}
        )
    
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF generation requires reportlab. Install with: pip install reportlab"
        )


async def generate_excel_report(request: ReportRequest):
    """Generate Excel format report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Fill, PatternFill, Alignment
        
        audits = filter_audits(request)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compliance Report"
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Headers
        headers = ['Audit ID', 'Product Name', 'Category', 'Compliance Score', 'Risk Level', 'Violations', 'Created At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, audit in enumerate(audits, 2):
            ws.cell(row=row, column=1, value=audit['audit_id'])
            ws.cell(row=row, column=2, value=audit['product_name'])
            ws.cell(row=row, column=3, value=audit['category'])
            ws.cell(row=row, column=4, value=audit['compliance_score'])
            ws.cell(row=row, column=5, value=audit['risk_level'])
            ws.cell(row=row, column=6, value=len(audit.get('violations', [])))
            ws.cell(row=row, column=7, value=audit['created_at'].strftime('%Y-%m-%d %H:%M'))
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=compliance_report.xlsx"}
        )
    
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel generation requires openpyxl. Install with: pip install openpyxl"
        )


def filter_audits(request: ReportRequest) -> list[dict]:
    """Filter audits based on request parameters."""
    audits = sample_audits.copy()
    
    if request.audit_ids:
        audits = [a for a in audits if a['audit_id'] in request.audit_ids]
    
    if request.date_from:
        audits = [a for a in audits if a['created_at'] >= request.date_from]
    
    if request.date_to:
        audits = [a for a in audits if a['created_at'] <= request.date_to]
    
    if request.category:
        audits = [a for a in audits if a['category'] == request.category]
    
    if request.risk_level:
        audits = [a for a in audits if a['risk_level'] == request.risk_level]
    
    return audits


def generate_summary(audits: list[dict]) -> dict:
    """Generate summary statistics."""
    if not audits:
        return {"total": 0, "message": "No audits found"}
    
    scores = [a['compliance_score'] for a in audits]
    
    return {
        "total_audits": len(audits),
        "average_score": round(sum(scores) / len(scores), 1),
        "min_score": min(scores),
        "max_score": max(scores),
        "compliant_count": sum(1 for a in audits if a['risk_level'] == 'compliant'),
        "high_risk_count": sum(1 for a in audits if a['risk_level'] in ['high', 'critical']),
        "total_violations": sum(len(a.get('violations', [])) for a in audits),
    }


def generate_violation_report(audits: list[dict]) -> dict:
    """Generate violation-focused report."""
    all_violations = []
    
    for audit in audits:
        for v in audit.get('violations', []):
            all_violations.append({
                "audit_id": audit['audit_id'],
                "product_name": audit['product_name'],
                "rule": v.get('rule'),
                "severity": v.get('severity'),
                "message": v.get('message'),
            })
    
    # Group by rule
    by_rule = {}
    for v in all_violations:
        rule = v.get('rule', 'unknown')
        if rule not in by_rule:
            by_rule[rule] = []
        by_rule[rule].append(v)
    
    return {
        "total_violations": len(all_violations),
        "unique_rules": len(by_rule),
        "violations_by_rule": {
            rule: {
                "count": len(violations),
                "examples": violations[:3]
            }
            for rule, violations in sorted(by_rule.items(), key=lambda x: -len(x[1]))
        },
        "all_violations": all_violations,
    }


@app.get("/reports/trends")
async def get_trends(
    days: int = Query(30, ge=1, le=365),
    metric: str = Query("score", description="Metric to trend: score, violations, audits"),
):
    """Get trend data for charts."""
    
    cutoff = datetime.utcnow() - timedelta(days=days)
    audits = [a for a in sample_audits if a['created_at'] >= cutoff]
    
    # Group by day
    daily_data = {}
    
    for audit in audits:
        day = audit['created_at'].strftime('%Y-%m-%d')
        if day not in daily_data:
            daily_data[day] = {'scores': [], 'violations': 0, 'count': 0}
        
        daily_data[day]['scores'].append(audit['compliance_score'])
        daily_data[day]['violations'] += len(audit.get('violations', []))
        daily_data[day]['count'] += 1
    
    # Calculate trends
    trend_data = []
    for day in sorted(daily_data.keys()):
        data = daily_data[day]
        avg_score = sum(data['scores']) / len(data['scores']) if data['scores'] else 0
        
        trend_data.append({
            "date": day,
            "average_score": round(avg_score, 1),
            "total_violations": data['violations'],
            "audit_count": data['count'],
        })
    
    return {
        "period_days": days,
        "data_points": len(trend_data),
        "trends": trend_data,
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8006)
